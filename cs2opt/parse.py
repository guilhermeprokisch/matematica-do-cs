"""Parse SlothSquadron's CS2 weapon spreadsheet into tidy CSVs.

Sheets used:
  - 'Raw Values'                    -> engine stats, one row per weapon (+Alt/scoped columns)
  - 'Firing Inaccuracy(Standing)'   -> cumulative total inaccuracy per bullet while spraying
  - 'Firing Inaccuracy(Crouching)'  -> same, crouched
  - 'Reload & Deploy Times'         -> reload (fire-ready) and deploy times

Verified relationships (against the sheet's own computed values):
  armored chest dmg(d)  = Damage * (WeaponArmorRatio/2) * RangeModifier^(d_units/500)
  accurate range        = 152.4 / total_inaccuracy  (i.e. spread radius in meters
                          at distance d = total_inaccuracy/1000 * d, "accurate" =
                          within a 1-foot-diameter circle)
  spray recurrence      I[k+1] = base + (I[k] - base + InaccuracyFire) * decay
                          (decay fitted per weapon; reproduces sheet curves to <1e-3)
"""
from __future__ import annotations

import math
from pathlib import Path

import openpyxl
import pandas as pd

UNIT_M = 0.0254  # 1 game unit = 1 inch

DISPLAY_NAMES = {
    "deagle": "Desert Eagle", "revolver": "R8 Revolver", "elite": "Dual Berettas",
    "fiveseven": "Five-SeveN", "glock": "Glock-18", "hkp2000": "P2000",
    "usp_silencer": "USP-S", "p250": "P250", "cz75a": "CZ75-Auto", "tec9": "Tec-9",
    "mag7": "MAG-7", "nova": "Nova", "sawedoff": "Sawed-Off", "xm1014": "XM1014",
    "bizon": "PP-Bizon", "mac10": "MAC-10", "mp7": "MP7", "mp5sd": "MP5-SD",
    "mp9": "MP9", "p90": "P90", "ump45": "UMP-45",
    "ak47": "AK-47", "aug": "AUG", "famas": "FAMAS", "galilar": "Galil AR",
    "m4a1": "M4A4", "m4a1_silencer": "M4A1-S", "sg556": "SG 553",
    "m249": "M249", "negev": "Negev",
    "awp": "AWP", "g3sg1": "G3SG1", "scar20": "SCAR-20", "ssg08": "SSG 08",
}

# Weapons whose real-world use is the scoped (Alt) mode.
SCOPED = {"awp", "g3sg1", "scar20", "ssg08"}
# Rifles with an optional scope; we model both modes and keep the better one.
OPTIONAL_SCOPE = {"aug", "sg556"}


def _f(v):
    try:
        if v is None or v == "":
            return math.nan
        if isinstance(v, str) and "&" in v:  # e.g. CZ75 '1.53 & 1.17'
            return max(float(p) for p in v.split("&"))
        return float(v)
    except (TypeError, ValueError):
        return math.nan


def load_raw(xlsx: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Raw Values"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    out, category = [], str(hdr[0])  # first block's section name lives in the header row
    for r in rows[1:]:
        if r[0] is None:
            continue
        if math.isnan(_f(r[1])):  # section header row repeats column titles
            category = str(r[0])
            continue
        rec = dict(zip(hdr, r))
        rec["engine_name"] = str(r[0])
        rec["category"] = category
        out.append(rec)
    df = pd.DataFrame(out)
    df["display_name"] = df["engine_name"].map(DISPLAY_NAMES)

    keep = {
        "WeaporArmorRatio": "armor_ratio", "Damage": "damage",
        "RangeModifier": "range_modifier", "HeadshotMultiplier": "hs_mult",
        "CycleTime": "cycle_time", "KillAward": "kill_award",
        "MaxPlayerSpeed": "move_speed", "clip_size": "mag_size",
        "WeaponPrice": "price", "Bullets": "pellets",
        "Spread": "spread", "InaccuracyCrouch": "inacc_crouch",
        "InaccuracyStand": "inacc_stand", "InaccuracyFire": "inacc_fire",
        "InaccuracyMove": "inacc_move", "InaccuracyJump": "inacc_jump",
        "InaccuracyMoveAlt": "inacc_move_alt", "InaccuracyJumpAlt": "inacc_jump_alt",
        "SpreadAlt": "spread_alt", "InaccuracyCrouchAlt": "inacc_crouch_alt",
        "InaccuracyStandAlt": "inacc_stand_alt", "InaccuracyFireAlt": "inacc_fire_alt",
        "CycleTimeAlt": "cycle_time_alt", "MaxPlayerSpeedAlt": "move_speed_alt",
        "FullAuto": "full_auto",
    }
    for src, dst in keep.items():
        df[dst] = df[src].map(_f)
    df["armor_pen"] = df["armor_ratio"] / 2.0
    cols = ["engine_name", "display_name", "category"] + list(keep.values()) + ["armor_pen"]
    return df[cols]


def load_spray_curves(xlsx: Path) -> pd.DataFrame:
    """Cumulative total inaccuracy (incl. Spread) per bullet, standing & crouching."""
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    frames = []
    for sheet, stance in [("Firing Inaccuracy(Standing)", "stand"),
                          ("Firing Inaccuracy(Crouching)", "crouch")]:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[1:]:
            if r[0] is None or r[1] is None:
                continue
            name = str(r[0])
            if name not in DISPLAY_NAMES:
                continue
            for k, v in enumerate(r[6:36], start=1):
                fv = _f(v)
                if not math.isnan(fv):
                    frames.append({"engine_name": name, "stance": stance,
                                   "bullet": k, "total_inaccuracy": fv})
    return pd.DataFrame(frames)


def load_reload(xlsx: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Reload & Deploy Times"]
    norm = {"".join(c for c in v.lower() if c.isalnum()): k
            for k, v in DISPLAY_NAMES.items()}
    norm["m4a4"] = "m4a1"  # display 'M4A4' -> engine 'm4a1'
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None or r[1] is None or "(" in str(r[0]):
            continue
        key = "".join(c for c in str(r[0]).lower() if c.isalnum())
        eng = norm.get(key)
        if eng:
            out.append({"engine_name": eng, "deploy_time": _f(r[1]),
                        "reload_clip_ready": _f(r[2]), "reload_fire_ready": _f(r[3])})
    return pd.DataFrame(out)


def fit_decay(curve: list[float], base: float, fire: float) -> float:
    """Fit per-cycle decay d in I[k+1] = base + (I[k]-base+fire)*d from a sheet curve.

    Only the rising portion is used: Negev/M249 curves fall again after their
    'laser' mechanic kicks in, which the constant-parameter recurrence can't
    represent.
    """
    if len(curve) < 2 or fire <= 0:
        return 0.0
    peak = max(range(len(curve)), key=lambda i: curve[i])
    rising = curve[: peak + 1]
    ds = []
    for a, b in zip(rising[:-1], rising[1:]):
        denom = (a - base) + fire
        if denom > 1e-9:
            ds.append(max(0.0, min(0.999, (b - base) / denom)))
    return sum(ds) / len(ds) if ds else 0.0


def build_bullet_curves(raw: pd.DataFrame, spray: pd.DataFrame,
                        max_bullets: int = 160) -> pd.DataFrame:
    """Per-weapon, per-mode, per-stance bullet inaccuracy sequences up to mag size.

    Sheet curves (primary mode) are used verbatim and extended by holding the
    last value (safe for Negev's special decreasing curve). Scoped (alt) mode
    curves are synthesized with the fitted decay and Alt base/fire values.
    """
    recs, params = [], []
    for _, w in raw.iterrows():
        name = w.engine_name
        mag = int(w.mag_size) if not math.isnan(w.mag_size) else 30
        for stance in ("stand", "crouch"):
            sheet_curve = (spray[(spray.engine_name == name) & (spray.stance == stance)]
                           .sort_values("bullet").total_inaccuracy.tolist())
            base = (w.inacc_stand if stance == "stand" else w.inacc_crouch) + w.spread
            fire = w.inacc_fire
            decay = fit_decay(sheet_curve, base, fire)

            # primary mode
            n = min(mag, max_bullets)
            curve = list(sheet_curve[:n])
            while len(curve) < n:
                curve.append(curve[-1] if curve else base)
            for k, v in enumerate(curve, start=1):
                recs.append({"engine_name": name, "mode": "primary", "stance": stance,
                             "bullet": k, "total_inaccuracy": v})
            params.append({"engine_name": name, "mode": "primary", "stance": stance,
                           "base_total": curve[0], "fire_inacc": fire,
                           "decay_per_cycle": decay})

            # scoped/alt mode
            if name in SCOPED | OPTIONAL_SCOPE and not math.isnan(w.inacc_stand_alt):
                base_a = ((w.inacc_stand_alt if stance == "stand" else w.inacc_crouch_alt)
                          + w.spread_alt)
                fire_a = w.inacc_fire_alt
                cur, alt = base_a, []
                for _k in range(n):
                    alt.append(cur)
                    cur = base_a + (cur - base_a + fire_a) * decay
                for k, v in enumerate(alt, start=1):
                    recs.append({"engine_name": name, "mode": "scoped", "stance": stance,
                                 "bullet": k, "total_inaccuracy": v})
                params.append({"engine_name": name, "mode": "scoped", "stance": stance,
                               "base_total": base_a, "fire_inacc": fire_a,
                               "decay_per_cycle": decay})
    return pd.DataFrame(recs), pd.DataFrame(params)


def main(xlsx: Path, outdir: Path) -> None:
    outdir.mkdir(parents=True, exist_ok=True)
    raw = load_raw(xlsx)
    spray = load_spray_curves(xlsx)
    reload_df = load_reload(xlsx)
    weapons = raw.merge(reload_df, on="engine_name", how="left")
    curves, spray_params = build_bullet_curves(raw, spray)
    weapons.to_csv(outdir / "weapons.csv", index=False)
    curves.to_csv(outdir / "bullet_curves.csv", index=False)
    spray_params.to_csv(outdir / "spray_params.csv", index=False)
    print(f"weapons: {len(weapons)} rows -> {outdir/'weapons.csv'}")
    print(f"curves:  {len(curves)} rows -> {outdir/'bullet_curves.csv'}")
    print(f"params:  {len(spray_params)} rows -> {outdir/'spray_params.csv'}")


if __name__ == "__main__":
    root = Path(__file__).resolve().parent.parent
    main(root / "data/source/weapon_spreadsheet.xlsx", root / "data")
